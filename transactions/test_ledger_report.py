"""
Tests for the admin finance ledger: filtering, totals, export and the failed-payments list.

The property that matters most here is agreement. The list, the summary card and the
downloaded spreadsheet all describe the same slice of money, and an operator reconciling a
disputed figure has to be able to trust that they match. Several tests below exist only to
pin that they are driven by one filter implementation rather than three.
"""

import csv
import io
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from rest_framework.test import APIClient

from transactions.ledger_report import (
    export_csv,
    export_xlsx,
    filtered_entries,
    summarise,
)
from transactions.models import LedgerEntry, PaystackEvent, Wallet

User = get_user_model()


class LedgerFilterTests(TestCase):

    def setUp(self):
        self.alice = User.objects.create_user(
            email='alice@test.com', password='test123', full_name='Alice A',
        )
        self.bob = User.objects.create_user(
            email='bob@test.com', password='test123', full_name='Bob B',
        )
        self.alice_wallet, _ = Wallet.objects.get_or_create(user=self.alice)
        self.bob_wallet, _ = Wallet.objects.get_or_create(user=self.bob)

        self.alice_wallet.credit(
            Decimal('5000.00'),
            source='Wallet deposit DEP-A',
            bucket=LedgerEntry.Bucket.SPENDABLE,
            entry_type=LedgerEntry.EntryType.DEPOSIT,
            reference='DEP-A',
        )
        self.alice_wallet.credit(
            Decimal('1000.00'),
            source='Referral bonus',
            entry_type=LedgerEntry.EntryType.REFERRAL_BONUS,
            reference='REF-A',
        )
        self.bob_wallet.credit(
            Decimal('3000.00'),
            source='Commission',
            entry_type=LedgerEntry.EntryType.COMMISSION,
            reference='COM-B',
        )
        self.bob_wallet.debit(
            Decimal('800.00'),
            source='Withdrawal WTH-B',
            bucket=LedgerEntry.Bucket.WITHDRAWABLE,
            entry_type=LedgerEntry.EntryType.WITHDRAWAL,
            reference='WTH-B',
        )

    def test_no_filters_returns_everything(self):
        self.assertEqual(filtered_entries({}).count(), 4)

    def test_filtering_by_user_email_is_exact_and_case_insensitive(self):
        self.assertEqual(filtered_entries({'user': 'alice@test.com'}).count(), 2)
        self.assertEqual(filtered_entries({'user': 'ALICE@TEST.COM'}).count(), 2)

    def test_filtering_by_direction(self):
        self.assertEqual(filtered_entries({'direction': 'DEBIT'}).count(), 1)
        self.assertEqual(filtered_entries({'direction': 'credit'}).count(), 3)

    def test_filtering_by_bucket(self):
        self.assertEqual(filtered_entries({'bucket': 'SPENDABLE'}).count(), 1)

    def test_filtering_by_several_entry_types_at_once(self):
        """One pass over a related group, e.g. a payout and its reversal."""
        qs = filtered_entries({'entry_type': 'WITHDRAWAL,DEPOSIT'})
        self.assertEqual(qs.count(), 2)

    def test_an_unknown_entry_type_is_ignored_rather_than_raising(self):
        """A bad value in an admin screen should widen the report, not 500 it."""
        self.assertEqual(filtered_entries({'entry_type': 'NOT_A_TYPE'}).count(), 4)

    def test_search_covers_reference_and_user(self):
        self.assertEqual(filtered_entries({'search': 'WTH-B'}).count(), 1)
        self.assertEqual(filtered_entries({'search': 'alice'}).count(), 2)

    def test_a_bare_date_to_includes_that_whole_day(self):
        """
        The most surprising way a finance report can be wrong: date_to=today excluding
        everything that happened today because a bare date reads as midnight.
        """
        today = timezone.localtime().strftime('%Y-%m-%d')
        self.assertEqual(filtered_entries({'date_to': today}).count(), 4)

    def test_a_date_from_after_everything_returns_nothing(self):
        future = (timezone.localtime() + timezone.timedelta(days=1)).strftime('%Y-%m-%d')
        self.assertEqual(filtered_entries({'date_from': future}).count(), 0)

    def test_an_unparseable_date_is_ignored(self):
        self.assertEqual(filtered_entries({'date_from': 'last tuesday'}).count(), 4)

    def test_filters_combine(self):
        qs = filtered_entries({'user': 'bob@test.com', 'direction': 'DEBIT'})
        self.assertEqual(qs.count(), 1)


class LedgerSummaryTests(TestCase):

    def setUp(self):
        self.user = User.objects.create_user(
            email='sum@test.com', password='test123', full_name='Sum User',
        )
        self.wallet, _ = Wallet.objects.get_or_create(user=self.user)
        self.wallet.credit(Decimal('5000.00'), source='Commission')
        self.wallet.credit(Decimal('2500.00'), source='Referral bonus')
        self.wallet.debit(
            Decimal('1500.00'),
            source='Withdrawal WTH-S',
            bucket=LedgerEntry.Bucket.WITHDRAWABLE,
        )

    def test_totals_and_net(self):
        summary = summarise(filtered_entries({}))

        self.assertEqual(summary['count'], 3)
        self.assertEqual(summary['total_credits'], '7500.00')
        self.assertEqual(summary['total_debits'], '1500.00')
        self.assertEqual(summary['net'], '6000.00')

    def test_an_empty_slice_reports_zeroes_rather_than_none(self):
        """A summary card must render something for a filter that matches nothing."""
        summary = summarise(filtered_entries({'user': 'nobody@test.com'}))

        self.assertEqual(summary['count'], 0)
        self.assertEqual(summary['total_credits'], '0.00')
        self.assertEqual(summary['net'], '0.00')

    def test_the_summary_respects_the_same_filters_as_the_list(self):
        qs = filtered_entries({'direction': 'DEBIT'})
        summary = summarise(qs)

        self.assertEqual(summary['count'], qs.count())
        self.assertEqual(summary['total_credits'], '0.00')
        self.assertEqual(summary['total_debits'], '1500.00')

    def test_breakdowns_split_by_type_and_bucket(self):
        summary = summarise(filtered_entries({}))

        self.assertTrue(any(row['direction'] == 'DEBIT' for row in summary['by_type']))
        bucket_names = {row['bucket'] for row in summary['by_bucket']}
        self.assertIn(LedgerEntry.Bucket.WITHDRAWABLE, bucket_names)


class LedgerExportTests(TestCase):

    def setUp(self):
        self.user = User.objects.create_user(
            email='export@test.com', password='test123', full_name='Export User',
        )
        self.wallet, _ = Wallet.objects.get_or_create(user=self.user)
        self.wallet.credit(
            Decimal('4000.00'),
            source='Commission',
            entry_type=LedgerEntry.EntryType.COMMISSION,
            reference='COM-X',
        )
        self.wallet.debit(
            Decimal('1000.00'),
            source='Withdrawal WTH-X',
            bucket=LedgerEntry.Bucket.WITHDRAWABLE,
            entry_type=LedgerEntry.EntryType.WITHDRAWAL,
            reference='WTH-X',
        )

    def _csv_rows(self, qs):
        response = export_csv(qs, 'test-export')
        body = b''.join(response.streaming_content).decode('utf-8')
        return list(csv.reader(io.StringIO(body)))

    def test_csv_has_a_header_and_a_row_per_entry(self):
        rows = self._csv_rows(filtered_entries({}))

        self.assertEqual(rows[0][0], 'Date')
        self.assertEqual(len([r for r in rows if r]), 3)  # header + 2 entries

    def test_csv_contains_only_the_filtered_rows(self):
        """The export must match what the operator was looking at, not the whole ledger."""
        rows = self._csv_rows(filtered_entries({'direction': 'DEBIT'}))

        data_rows = [r for r in rows[1:] if r]
        self.assertEqual(len(data_rows), 1)
        self.assertIn('WTH-X', ','.join(data_rows[0]))

    def test_debits_export_with_a_negative_signed_amount(self):
        """So a column of signed amounts sums to the net movement."""
        rows = self._csv_rows(filtered_entries({'direction': 'DEBIT'}))
        header = rows[0]
        signed = rows[1][header.index('Signed Amount (NGN)')]

        self.assertTrue(signed.startswith('-'), signed)

    def test_credits_export_with_a_positive_signed_amount(self):
        rows = self._csv_rows(filtered_entries({'direction': 'CREDIT'}))
        header = rows[0]
        signed = rows[1][header.index('Signed Amount (NGN)')]

        self.assertFalse(signed.startswith('-'), signed)

    def test_the_export_names_the_wallet_owner(self):
        rows = self._csv_rows(filtered_entries({}))
        self.assertIn('export@test.com', ','.join(rows[1]))

    def test_an_empty_export_still_has_its_header(self):
        rows = self._csv_rows(filtered_entries({'user': 'nobody@test.com'}))
        self.assertEqual(rows[0][0], 'Date')
        self.assertEqual(len([r for r in rows[1:] if r]), 0)

    def test_xlsx_writes_amounts_as_numbers_not_text(self):
        """The reason to offer XLSX at all: the amounts have to sum in the sheet."""
        from openpyxl import load_workbook

        response = export_xlsx(filtered_entries({}), 'test-export')
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook['Ledger']

        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        amount_col = header.index('Amount (NGN)')
        first_data_row = next(sheet.iter_rows(min_row=2, max_row=2))

        self.assertIsInstance(first_data_row[amount_col].value, (int, float))

    def test_xlsx_reports_truncation_rather_than_silently_dropping_rows(self):
        response = export_xlsx(filtered_entries({}), 'test-export', limit=1)

        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(response.content))
        text = ' '.join(
            str(cell.value)
            for row in workbook['Ledger'].iter_rows()
            for cell in row
            if cell.value
        )
        self.assertIn('Truncated', text)


class AdminLedgerEndpointTests(TestCase):

    def setUp(self):
        self.admin = User.objects.create_user(
            email='ledgeradmin@test.com',
            password='test123',
            full_name='Ledger Admin',
            role=User.Role.BUSINESS_ADMIN,
        )
        self.customer = User.objects.create_user(
            email='ledgercust@test.com', password='test123', full_name='Ledger Cust',
        )
        wallet, _ = Wallet.objects.get_or_create(user=self.customer)
        wallet.credit(Decimal('2000.00'), source='Commission')
        wallet.debit(
            Decimal('500.00'),
            source='Withdrawal WTH-E',
            bucket=LedgerEntry.Bucket.WITHDRAWABLE,
        )

        self.client = APIClient()
        self.client.force_authenticate(user=self.admin)

    def test_ledger_list_is_admin_only(self):
        self.client.force_authenticate(user=self.customer)

        response = self.client.get('/transactions/admin/ledger/')

        self.assertIn(response.status_code, (401, 403))

    def test_export_is_admin_only(self):
        """The export is a separate view, so it needs its own check - not just the list."""
        self.client.force_authenticate(user=self.customer)

        response = self.client.get('/transactions/admin/ledger/export/')

        self.assertIn(response.status_code, (401, 403))

    def test_summary_is_admin_only(self):
        self.client.force_authenticate(user=self.customer)

        response = self.client.get('/transactions/admin/ledger/summary/')

        self.assertIn(response.status_code, (401, 403))

    def test_admin_sees_the_ledger(self):
        response = self.client.get('/transactions/admin/ledger/')

        self.assertEqual(response.status_code, 200)
        results = response.data['results'] if isinstance(response.data, dict) else response.data
        self.assertEqual(len(results), 2)

    def test_the_summary_matches_the_list_for_the_same_filters(self):
        list_response = self.client.get('/transactions/admin/ledger/?direction=DEBIT')
        summary_response = self.client.get(
            '/transactions/admin/ledger/summary/?direction=DEBIT'
        )

        results = (
            list_response.data['results']
            if isinstance(list_response.data, dict)
            else list_response.data
        )
        self.assertEqual(summary_response.data['data']['count'], len(results))
        self.assertEqual(summary_response.data['data']['total_debits'], '500.00')

    def test_the_summary_echoes_the_filters_it_applied(self):
        response = self.client.get('/transactions/admin/ledger/summary/?direction=DEBIT')

        self.assertEqual(response.data['data']['filters']['direction'], 'DEBIT')

    def test_csv_export_downloads_as_an_attachment(self):
        response = self.client.get('/transactions/admin/ledger/export/?export_format=csv')

        self.assertEqual(response.status_code, 200)
        self.assertIn('text/csv', response['Content-Type'])
        self.assertIn('attachment;', response['Content-Disposition'])

    def test_xlsx_export_downloads_as_a_spreadsheet(self):
        response = self.client.get('/transactions/admin/ledger/export/?export_format=xlsx')

        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])

    def test_an_unknown_export_format_is_refused(self):
        response = self.client.get('/transactions/admin/ledger/export/?export_format=pdf')

        self.assertEqual(response.status_code, 400)

    def test_the_export_honours_the_list_filters(self):
        response = self.client.get(
            '/transactions/admin/ledger/export/?export_format=csv&direction=DEBIT'
        )
        body = b''.join(response.streaming_content).decode('utf-8')
        rows = [r for r in csv.reader(io.StringIO(body)) if r]

        self.assertEqual(len(rows), 2)  # header + the single debit


class AdminFailedPaymentsTests(TestCase):

    def setUp(self):
        self.admin = User.objects.create_user(
            email='failadmin@test.com',
            password='test123',
            full_name='Fail Admin',
            role=User.Role.BUSINESS_ADMIN,
        )
        PaystackEvent.objects.create(
            event_id='charge.success:1', event_type='charge.success',
            reference='ORD-1', status=PaystackEvent.Status.PROCESSED,
        )
        PaystackEvent.objects.create(
            event_id='charge.success:2', event_type='charge.success',
            reference='ORD-2', status=PaystackEvent.Status.FAILED,
            error_message='Payment not found',
        )
        PaystackEvent.objects.create(
            event_id='refund.processed:3', event_type='refund.processed',
            reference='DEP-3', status=PaystackEvent.Status.IGNORED,
            error_message='no wallet deposit for refunded transaction DEP-3',
        )

        self.client = APIClient()
        self.client.force_authenticate(user=self.admin)

    def _results(self, response):
        return response.data['results'] if isinstance(response.data, dict) else response.data

    def test_it_defaults_to_the_states_worth_attention(self):
        """PROCESSED events are the ledger seen from the other side, not a problem."""
        response = self.client.get('/transactions/admin/failed-payments/')

        self.assertEqual(response.status_code, 200)
        results = self._results(response)
        self.assertEqual(len(results), 2)
        self.assertNotIn('PROCESSED', [r['status'] for r in results])

    def test_it_carries_the_reason_so_an_operator_can_act(self):
        response = self.client.get('/transactions/admin/failed-payments/')
        results = self._results(response)

        self.assertTrue(any(r['error_message'] for r in results))

    def test_filtering_by_status(self):
        response = self.client.get('/transactions/admin/failed-payments/?status=FAILED')

        self.assertEqual(len(self._results(response)), 1)

    def test_filtering_by_event_type(self):
        response = self.client.get(
            '/transactions/admin/failed-payments/?event_type=refund'
        )

        self.assertEqual(len(self._results(response)), 1)

    def test_failed_events_never_appear_in_the_ledger(self):
        """
        The separation this whole screen exists for. A failed webhook moved no money, so
        including it in the ledger would make every finance total wrong.
        """
        self.assertEqual(filtered_entries({}).count(), 0)

    def test_it_is_admin_only(self):
        customer = User.objects.create_user(
            email='nosy@test.com', password='test123', full_name='Nosy',
        )
        self.client.force_authenticate(user=customer)

        response = self.client.get('/transactions/admin/failed-payments/')

        self.assertIn(response.status_code, (401, 403))

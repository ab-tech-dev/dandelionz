"""
The admin finance ledger: filtering, totals, and export.

One filter definition serves the list, the summary and the export. That is the point of
this module: an operator reconciling a figure has to be able to trust that the CSV they
downloaded contains exactly the rows they were looking at. Three separate filter
implementations would drift, and the first sign of it would be a total that does not match
its own export.

The ledger is append-only, so everything here is read-only by construction.
"""

import csv
from datetime import datetime, time
from decimal import Decimal

from django.db.models import Count, Q, Sum
from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone

from transactions.models import LedgerEntry, money

# Columns shared by both export formats, in the order an accountant reads them: when, who,
# what, how much, and then the identifiers needed to chase a row back to its source.
EXPORT_COLUMNS = [
    ('created_at', 'Date'),
    ('user_email', 'User'),
    ('entry_type', 'Type'),
    ('direction', 'Direction'),
    ('bucket', 'Bucket'),
    ('amount', 'Amount (NGN)'),
    ('signed_amount', 'Signed Amount (NGN)'),
    ('balance_after', 'Balance After (NGN)'),
    ('reference', 'Reference'),
    ('description', 'Description'),
    ('order_id', 'Order'),
    ('payout_reference', 'Payout'),
    ('operation_key', 'Operation Key'),
]


DATE_ONLY_FORMAT = '%Y-%m-%d'
_DATETIME_FORMATS = ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', DATE_ONLY_FORMAT)


def _parse_date(value, end_of_day=False):
    """
    Accept either a date or a full timestamp.

    A bare date is read as local midnight, and `end_of_day` pushes it to 23:59:59.999999 -
    otherwise `date_to=2026-07-20` would silently exclude everything that happened on the
    20th, which is the single most surprising way a finance report can be wrong.

    That promotion keys off which *format matched*, not off the parsed time being midnight.
    Testing `hour == 0 and minute == 0` cannot tell a bare date from a deliberate
    `2026-07-21 00:00:00`, so an operator asking for a precise cutoff would silently get a
    whole extra day folded into the total they were about to sign off.
    """
    if not value:
        return None

    text = str(value).strip()
    parsed = None
    matched_format = None
    for fmt in _DATETIME_FORMATS:
        try:
            parsed = datetime.strptime(text, fmt)
            matched_format = fmt
            break
        except ValueError:
            continue

    if parsed is None:
        return None

    if end_of_day and matched_format == DATE_ONLY_FORMAT:
        parsed = datetime.combine(parsed.date(), time.max)

    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def parse_filters(params):
    """
    Normalise raw query parameters once, so every consumer sees the same values.

    filtered_entries() and applied_filters() both need the parsed dates. Parsing in each
    of them separately meant the "filters applied" block echoed back to the operator was
    computed independently of the query that actually ran - so a change to one could leave
    the screen describing a range the rows did not come from.
    """
    entry_type = params.get('entry_type')
    types = []
    if entry_type:
        wanted = [t.strip().upper() for t in str(entry_type).split(',') if t.strip()]
        types = [t for t in wanted if t in LedgerEntry.EntryType.values]

    direction = params.get('direction')
    direction = str(direction).upper() if direction else ''

    bucket = params.get('bucket')
    bucket = str(bucket).upper() if bucket else ''

    def _clean(name):
        value = params.get(name)
        return str(value).strip() if value else ''

    return {
        'date_from': _parse_date(params.get('date_from')),
        'date_to': _parse_date(params.get('date_to'), end_of_day=True),
        'entry_types': types,
        'direction': direction if direction in LedgerEntry.Direction.values else '',
        'bucket': bucket if bucket in LedgerEntry.Bucket.values else '',
        'user': _clean('user'),
        'reference': _clean('reference'),
        'search': _clean('search'),
    }


def filtered_entries(params):
    """
    Build the ledger queryset for a set of query parameters.

    `params` is any mapping - request.query_params in a view, a plain dict in a test.
    Unrecognised or unparseable values are ignored rather than raising: a report that
    quietly widens is better than a 500 in an admin screen, and the applied filters are
    echoed back in the response so an operator can see what actually ran.
    """
    f = parse_filters(params)

    qs = (
        LedgerEntry.objects
        .select_related('wallet', 'wallet__user', 'order', 'payout_request')
        .order_by('-created_at', '-id')
    )

    if f['date_from']:
        qs = qs.filter(created_at__gte=f['date_from'])

    if f['date_to']:
        qs = qs.filter(created_at__lte=f['date_to'])

    # Comma-separated so a report can cover a related group in one pass, e.g.
    # WITHDRAWAL,WITHDRAWAL_REVERSAL to see a payout and its reversal together.
    if f['entry_types']:
        qs = qs.filter(entry_type__in=f['entry_types'])

    if f['direction']:
        qs = qs.filter(direction=f['direction'])

    if f['bucket']:
        qs = qs.filter(bucket=f['bucket'])

    if f['user']:
        qs = qs.filter(wallet__user__email__iexact=f['user'])

    if f['reference']:
        qs = qs.filter(reference__icontains=f['reference'])

    if f['search']:
        term = f['search']
        qs = qs.filter(
            Q(reference__icontains=term)
            | Q(description__icontains=term)
            | Q(wallet__user__email__icontains=term)
            | Q(wallet__user__full_name__icontains=term)
        )

    return qs


def applied_filters(params):
    """
    Echo back what was actually applied, so a screen can show it and an export can say so.

    Reports the *normalised* values rather than the raw input: an entry_type the backend
    rejected as unknown should not be echoed as though it narrowed anything.
    """
    f = parse_filters(params)
    return {
        'date_from': f['date_from'].isoformat() if f['date_from'] else None,
        'date_to': f['date_to'].isoformat() if f['date_to'] else None,
        'entry_type': ','.join(f['entry_types']) if f['entry_types'] else None,
        'direction': f['direction'] or None,
        'bucket': f['bucket'] or None,
        'user': f['user'] or None,
        'reference': f['reference'] or None,
        'search': f['search'] or None,
    }


def summarise(qs):
    """
    Totals for a filtered ledger queryset.

    `net` is credits minus debits: what the filtered slice did to the platform's balances
    overall. It is not a balance - the ledger spans every wallet - which is why the two
    directions are also reported separately rather than only as a single number.
    """
    totals = qs.aggregate(
        credits=Sum('amount', filter=Q(direction=LedgerEntry.Direction.CREDIT)),
        debits=Sum('amount', filter=Q(direction=LedgerEntry.Direction.DEBIT)),
        count=Count('id'),
    )
    # Sum() returns an unquantized Decimal, so a total of 7500 serialises as "7500" while
    # every individual amount carries two places. Money is always rendered to 2dp.
    credits = money(totals['credits'] or Decimal('0.00'))
    debits = money(totals['debits'] or Decimal('0.00'))

    by_type = list(
        qs.values('entry_type', 'direction')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('entry_type', 'direction')
    )

    by_bucket = list(
        qs.values('bucket', 'direction')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('bucket', 'direction')
    )

    return {
        'count': totals['count'] or 0,
        'total_credits': str(credits),
        'total_debits': str(debits),
        'net': str(money(credits - debits)),
        'by_type': [
            {
                'entry_type': row['entry_type'],
                'direction': row['direction'],
                'total': str(money(row['total'] or Decimal('0.00'))),
                'count': row['count'],
            }
            for row in by_type
        ],
        'by_bucket': [
            {
                'bucket': row['bucket'],
                'direction': row['direction'],
                'total': str(money(row['total'] or Decimal('0.00'))),
                'count': row['count'],
            }
            for row in by_bucket
        ],
    }


def _row_values(entry):
    """One ledger entry as the flat values both export formats write."""
    signed = entry.amount if entry.direction == LedgerEntry.Direction.CREDIT else -entry.amount
    return {
        'created_at': timezone.localtime(entry.created_at).strftime('%Y-%m-%d %H:%M:%S'),
        # wallet is a non-nullable PROTECT FK, so there is always an owner to name.
        'user_email': entry.wallet.user.email,
        'entry_type': entry.entry_type,
        'direction': entry.direction,
        'bucket': entry.bucket,
        'amount': entry.amount,
        'signed_amount': signed,
        'balance_after': entry.balance_after,
        'reference': entry.reference or '',
        'description': entry.description or '',
        'order_id': str(entry.order.order_id) if entry.order_id else '',
        'payout_reference': entry.payout_request.reference if entry.payout_request_id else '',
        'operation_key': entry.operation_key or '',
    }


class _Echo:
    """A file-like object whose write() returns the line, for streaming csv output."""

    def write(self, value):
        return value


def export_csv(qs, filename):
    """
    Stream the filtered ledger as CSV.

    Streamed rather than built in memory: a finance export is exactly the request most
    likely to cover a year, and the ledger has a row for every movement of money on the
    platform. `iterator()` keeps the queryset from being cached whole at the same time.
    """
    writer = csv.writer(_Echo())
    headers = [label for _, label in EXPORT_COLUMNS]

    def rows():
        yield writer.writerow(headers)
        for entry in qs.iterator(chunk_size=2000):
            values = _row_values(entry)
            yield writer.writerow([values[key] for key, _ in EXPORT_COLUMNS])

    response = StreamingHttpResponse(rows(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    return response


def export_xlsx(qs, filename, limit=100000):
    """
    Build the filtered ledger as a real spreadsheet.

    Unlike CSV this cannot stream - the format is a zip archive finalised at the end - so it
    is capped. write_only mode keeps memory flat per row; the cap exists because the
    response still has to be assembled in full before it is sent, and an uncapped export of
    a large ledger would take the worker down rather than merely being slow.

    Amounts are written as numbers, not strings, so they sum in the sheet. That is the whole
    reason to offer XLSX alongside CSV.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet('Ledger')

    numeric = {'amount', 'signed_amount', 'balance_after'}
    header_font = Font(bold=True)

    from openpyxl.cell import WriteOnlyCell

    header_cells = []
    for _, label in EXPORT_COLUMNS:
        cell = WriteOnlyCell(sheet, value=label)
        cell.font = header_font
        header_cells.append(cell)
    sheet.append(header_cells)

    sheet.column_dimensions[get_column_letter(1)].width = 20
    sheet.column_dimensions[get_column_letter(2)].width = 28

    truncated = False
    for index, entry in enumerate(qs.iterator(chunk_size=2000)):
        if index >= limit:
            truncated = True
            break
        values = _row_values(entry)
        sheet.append([
            float(values[key]) if key in numeric else values[key]
            for key, _ in EXPORT_COLUMNS
        ])

    if truncated:
        sheet.append([])
        sheet.append([
            f"Truncated at {limit:,} rows. Narrow the date range, or use CSV for the "
            f"complete export."
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    workbook.save(response)
    return response
